"""Parse the Human Capital 'Workflow Approval' workbook into structured records.

Each per-unit sheet (e.g. IMU-CS, IMN-SBY) is a purchasing authority matrix:
a header (entity + project/unit), a set of nominal threshold bands as columns,
a set of authority-level rows with a Y/N per band, then Buyer / PR Creator, and
finally a Service Entry Sheet (SES) block with a creator/approver per band.

Column positions drift between sheets, so everything is located by its label
text rather than by fixed coordinates.
"""

import json
import re
import sys

import openpyxl


def norm(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def find_cell(ws, needle, max_row=40, max_col=25):
    """Return (row, col) of the first cell whose text contains needle."""
    needle = needle.lower()
    for r in range(1, min(ws.max_row, max_row) + 1):
        for c in range(1, min(ws.max_column, max_col) + 1):
            if needle in norm(ws.cell(r, c).value).lower():
                return r, c
    return None


def value_right_of(ws, row, col, span=8):
    """First non-empty cell to the right of (row, col)."""
    for c in range(col + 1, col + 1 + span):
        v = norm(ws.cell(row, c).value)
        if v:
            return v, c
    return "", None


def band_columns(ws):
    """Locate the threshold-band header row and its columns.

    Bands look like '< US$1k or <Rp.16,000,000'. Returns (row, [(col, label)]).
    """
    for r in range(1, min(ws.max_row, 20) + 1):
        cols = []
        for c in range(1, min(ws.max_column, 25) + 1):
            v = norm(ws.cell(r, c).value)
            if re.search(r"US\$?\s?\d|Rp\.?\s?\d", v) and ("<" in v or "–" in v or "-" in v or "≥" in v):
                cols.append((c, v))
        # A real band row has several of these side by side.
        if len(cols) >= 4:
            return r, cols
    return None, []


def parse_unit_sheet(ws):
    entity_hit = find_cell(ws, "COMPANY NAME") or find_cell(ws, "NAMA ENTITAS")
    project_hit = find_cell(ws, "PROYEK") or find_cell(ws, "PROJECT")
    if not entity_hit:
        return None

    entity_name, _ = value_right_of(ws, *entity_hit)
    project_name = ""
    if project_hit:
        project_name, _ = value_right_of(ws, *project_hit)

    band_row, bands = band_columns(ws)
    if not bands:
        return None
    band_cols = [c for c, _ in bands]

    # Locate the matrix header row: "Level / Tingkat", "Name / Nama", ...
    level_hit = find_cell(ws, "Level / Tingkat") or find_cell(ws, "Level")
    name_hit = find_cell(ws, "Name / Nama") or find_cell(ws, "Nama")
    pos_hit = find_cell(ws, "Position")
    comment_hit = find_cell(ws, "Comment") or find_cell(ws, "Komentar")
    if not (level_hit and name_hit):
        return None

    level_col = level_hit[1]
    name_col = name_hit[1]
    pos_col = pos_hit[1] if pos_hit else None
    comment_col = comment_hit[1] if comment_hit else None
    header_row = level_hit[0]

    def cell_after(row, col, span=6):
        """Merged 'name' cells often start a couple columns off; scan forward."""
        for c in range(col, col + span):
            v = norm(ws.cell(row, c).value)
            if v:
                return v
        return ""

    # Bound each scan so it stops before the next labelled column instead of
    # bleeding into it (e.g. an empty Name cell must not pick up a stray
    # value sitting in the Position column further right).
    name_span = max(1, pos_col - name_col) if pos_col else 6
    level_span = max(1, name_col - level_col)

    rows = []
    ses = []
    r = header_row + 1
    # Walk down until we hit the SES section or run out of rows.
    while r <= ws.max_row:
        level = norm(ws.cell(r, level_col).value)
        if level.lower().startswith("service entry"):
            break
        # Y/N matrix cells for this row
        yn = {}
        any_yn = False
        for c in band_cols:
            v = norm(ws.cell(r, c).value).upper()
            if v in ("Y", "N"):
                yn[c] = v == "Y"
                any_yn = True
        name = cell_after(r, name_col, name_span)
        position = norm(ws.cell(r, pos_col).value) if pos_col else ""
        comment = norm(ws.cell(r, comment_col).value) if comment_col else ""
        if position in ("0", "#N/A"):
            position = ""

        kind = None
        low = level.lower()
        if any_yn and level:
            kind = "authority"
        elif low.startswith("buyer"):
            kind = "buyer"
        elif low.startswith("purchase request creator") or low.startswith("pr creator"):
            kind = "creator"

        if kind == "authority":
            rows.append({
                "kind": "authority", "level": level, "name": name,
                "position": position, "comment": comment,
                "bands": {band_cols.index(c): yn[c] for c in yn},
            })
        elif kind in ("buyer", "creator"):
            # The person's name usually sits on the next row for these blocks.
            person = name
            if not person:
                person = cell_after(r + 1, name_col, name_span) or cell_after(r + 1, level_col, level_span)
            rows.append({"kind": kind, "level": level, "name": person,
                         "position": position, "comment": comment, "bands": {}})
        r += 1

    # SES block: find its own band-labelled rows plus creator/approver columns.
    ses_hdr = find_cell(ws, "SES Creator")
    ses_appr = find_cell(ws, "SES Approver")
    if ses_hdr:
        creator_col = ses_hdr[1]
        approver_col = ses_appr[1] if ses_appr else creator_col + 3
        start = ses_hdr[0] + 1
        for rr in range(start, ws.max_row + 1):
            # band label is in the far-left region of the SES block
            label = ""
            for c in range(1, creator_col):
                v = norm(ws.cell(rr, c).value)
                if re.search(r"US\$?\s?\d|Rp", v):
                    label = v
                    break
            if not label:
                continue
            creator = cell_after(rr, creator_col)
            approver = cell_after(rr, approver_col)
            ses.append({"band_label": label, "creator": creator, "approver": approver})

    return {
        "entity_name": entity_name,
        "project_name": project_name,
        "bands": [norm(lbl) for _, lbl in bands],
        "rows": rows,
        "ses": ses,
    }


def main():
    path = sys.argv[1]
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    skipped = []
    for ws in wb.worksheets:
        title = ws.title
        if title == "Sheet1" or "not use" in title.lower() or title == "IMU":
            skipped.append(title)
            continue
        parsed = parse_unit_sheet(ws)
        if parsed is None:
            skipped.append(title)
            continue
        out[title] = parsed
    print(json.dumps({"units": out, "skipped": skipped}, ensure_ascii=False, indent=1))


if __name__ == "__main__":
    main()
