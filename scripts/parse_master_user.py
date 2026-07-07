"""Parse the 'Master User (Lengkap)' workbook into structured records.

Layout: one flat table, columns = Entity, Site/Lokasi, Department, Role,
Nama, Email. Two distinct row shapes appear within each
(entity, site, department, role) block, in original row order:

- Roster roles (PR Creator, SES Creator, Buyer, Cost Reviewer): every row is
  one independent person with their own email.
- Hierarchy roles (Assignment Director, Deputy Director, Director, Division
  Head, President Director): each person is written as TWO consecutive rows
  with no email -- a Name row followed by a Position (job title) row.

A known source glitch splits a literal "#N/A" position into two stray rows
("#N", "A"); this is normalized back into one "#N/A" row before pairing.
"""

import json
import os
import sys

import openpyxl

ROSTER_ROLES = {"PR Creator", "SES Creator", "Buyer", "Cost Reviewer"}

_KNOWN_POSITIONS_PATH = os.path.join(os.path.dirname(__file__), "known_positions.json")
with open(_KNOWN_POSITIONS_PATH) as _f:
    KNOWN_POSITIONS = set(json.load(_f))


def norm(v):
    return v.strip() if isinstance(v, str) else v


VALID_ENTITIES = {"IMU", "ILSS", "IMN", "KGTE", "IRB", "PGE", "MBN", "ISB", "INDIS"}


def load_rows(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["1. Master User (Lengkap)"]
    rows = []
    for r in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        entity, site, dept, role, nama, email = (norm(v) for v in r)
        if entity is None:
            continue
        if entity not in VALID_ENTITIES:
            # Trailing color-legend block ("KETERANGAN WARNA / COLOR LEGEND"
            # and its entries) uses the Entity column for unrelated text --
            # once real entity rows stop, nothing after is data.
            break
        rows.append({"entity": entity, "site": site, "department": dept,
                      "role": role, "nama": nama, "email": email})
    return rows


def _degap_hash_na(rows):
    """Merge the stray '#N' + 'A' row pair back into a single '#N/A' row."""
    out = []
    i = 0
    while i < len(rows):
        cur = rows[i]
        if (cur["nama"] == "#N" and i + 1 < len(rows)
                and rows[i + 1]["nama"] == "A"
                and rows[i + 1]["role"] == cur["role"]):
            merged = dict(cur)
            merged["nama"] = "#N/A"
            out.append(merged)
            i += 2
        else:
            out.append(cur)
            i += 1
    return out


def parse(path):
    """Return (workspaces, roster_assignments, hierarchy_assignments).

    workspaces: list of dicts {entity, site, department} (deduped, ordered).
    roster_assignments: list of {entity, site, department, role, nama, email}.
    hierarchy_assignments: list of {entity, site, department, role, nama, position}.
    """
    rows = load_rows(path)

    workspaces = []
    seen_ws = set()
    roster = []
    hierarchy = []

    # Group consecutive rows sharing the same (entity, site, department, role),
    # preserving file order (the source is already grouped this way).
    i = 0
    n = len(rows)
    while i < n:
        r = rows[i]
        key = (r["entity"], r["site"], r["department"])
        if key not in seen_ws:
            seen_ws.add(key)
            workspaces.append({"entity": r["entity"], "site": r["site"], "department": r["department"]})

        block_key = (r["entity"], r["site"], r["department"], r["role"])
        j = i
        block = []
        while j < n and (rows[j]["entity"], rows[j]["site"], rows[j]["department"], rows[j]["role"]) == block_key:
            block.append(rows[j])
            j += 1

        if r["role"] in ROSTER_ROLES:
            for row in block:
                roster.append({
                    "entity": row["entity"], "site": row["site"], "department": row["department"],
                    "role": row["role"], "nama": row["nama"], "email": row["email"],
                })
        else:
            # Rows alternate between one-or-more consecutive person-name rows
            # and a single shared position row (two people can share one role
            # + position, written as two name rows then one position row).
            # A trailing name row with no position row after it just has no
            # recorded position.
            block = _degap_hash_na(block)
            pending_names = []
            for row in block:
                if row["nama"] in KNOWN_POSITIONS or row["nama"] == "#N/A":
                    for name_row in pending_names:
                        hierarchy.append({
                            "entity": name_row["entity"], "site": name_row["site"],
                            "department": name_row["department"], "role": name_row["role"],
                            "nama": name_row["nama"], "position": row["nama"],
                        })
                    pending_names = []
                else:
                    pending_names.append(row)
            for name_row in pending_names:
                hierarchy.append({
                    "entity": name_row["entity"], "site": name_row["site"],
                    "department": name_row["department"], "role": name_row["role"],
                    "nama": name_row["nama"], "position": None,
                })
        i = j

    return workspaces, roster, hierarchy


if __name__ == "__main__":
    ws, roster, hier = parse(sys.argv[1])
    print("workspaces:", len(ws))
    print("roster assignments:", len(roster))
    print("hierarchy assignments:", len(hier))
