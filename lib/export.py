import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for row in rows:
        ws.append(row)

    ws.freeze_panes = "A2"
    for col_idx, header in enumerate(headers, start=1):
        cell_lengths = [len(str(row[col_idx - 1])) if row[col_idx - 1] is not None else 0 for row in rows]
        width = max([len(str(header))] + cell_lengths) + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 50)

    return ws


def build_export_workbook(cur):
    wb = Workbook()
    wb.remove(wb.active)

    cur.execute("SELECT code, name FROM entities ORDER BY code")
    entities = cur.fetchall()
    _write_sheet(
        wb, "Entities", ["Code", "Name"],
        [[e["code"], e["name"]] for e in entities],
    )

    cur.execute(
        """
        SELECT e.code AS entity_code, s.name
        FROM sites s
        JOIN entities e ON e.id = s.entity_id
        ORDER BY e.code, s.name
        """
    )
    sites = cur.fetchall()
    _write_sheet(
        wb, "Sites", ["Entity", "Site Name"],
        [[s["entity_code"], s["name"]] for s in sites],
    )

    cur.execute("SELECT name, is_centralized FROM roles ORDER BY name")
    roles = cur.fetchall()
    _write_sheet(
        wb, "Roles", ["Role", "Centralized"],
        [[r["name"], "Yes" if r["is_centralized"] else "No"] for r in roles],
    )

    cur.execute(
        """
        SELECT e.code AS entity_code, s.name AS site_name, r.name AS role_name,
               ua.full_name, ua.email, ua.is_active
        FROM user_assignments ua
        JOIN entities e ON e.id = ua.entity_id
        JOIN roles r ON r.id = ua.role_id
        LEFT JOIN sites s ON s.id = ua.site_id
        ORDER BY e.code, r.name, ua.full_name
        """
    )
    assignments = cur.fetchall()
    _write_sheet(
        wb, "Assignments", ["Entity", "Site", "Role", "Full Name", "Email", "Active"],
        [
            [a["entity_code"], a["site_name"] or "-", a["role_name"], a["full_name"], a["email"],
             "Yes" if a["is_active"] else "No"]
            for a in assignments
        ],
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
